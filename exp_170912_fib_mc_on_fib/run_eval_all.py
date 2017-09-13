
import numpy as np
import os

import sys
sys.path.append(
    '/home/jhennies/src/nature_methods_multicut_pipeline/nature_methods_multicut_pipeline/software/')


def error_measure_results_to_xlsx_file(filepath, results, thresh_range, ds_names,
                                       row_offset=0, append=False, title=None):

    def write_list(ws, r, c, l):
        l = l.squeeze()
        assert l.ndim == 1
        # FIXME Is there a simpler command ommiting the for loop to write a list?
        for id, item in enumerate(l):
            ws.cell(row=r + id, column=c).value = item

    from openpyxl import Workbook, load_workbook

    if append:
        workbook = load_workbook(filepath)
    else:
        workbook = Workbook()

    worksheet = workbook.active

    if title is not None:
        worksheet.cell(row=1 + row_offset, column=1).value=title
        row_offset += 1

    # worksheet.cell(row=3, column=1).value = thresh_range
    write_list(worksheet, row_offset + 3, 1, thresh_range)

    for ds_id, ds_name in enumerate(ds_names):

        spl_str = '{}'.format(ds_name)

        worksheet.cell(row=1 + row_offset, column=2 + ds_id * (len(results.keys()) + 1)).value = spl_str

        for measure_id, measure in enumerate(results.keys()):

            col = 2 + ds_id * (len(results.keys()) + 1) + measure_id

            worksheet.cell(row=2 + row_offset, column=col).value = measure
            write_list(worksheet, 3 + row_offset, col, results[measure][ds_id])

    workbook.save(filepath)


def run_roi_and_rand_general():

    from eval_all import roi_and_rand_general

    # project_folder = '/mnt/localdata0/jhennies/results/multicut_workflow/170912_fib_mc_on_fib/'
    project_folder = '/mnt/localdata1/jhennies/neuraldata/results/'
    # result_files = ['result.h5']
    result_files = ['fib_7_5_7_neuroproof-fib25seg-med.h5']
    result_keys = ['data']
    # result_keys = ['beta_0.5']

    ds_names = ['fib_mc_constantin']
    # ds_names = ['fib_7_5_7', 'fib_8_5_6', 'fib_8_5_7']
    # ds_names = ['fib_8_5_6', 'fib_8_5_7']

    import vigra
    gt = vigra.readHDF5('/mnt/localdata1/jhennies/neuraldata/results/fib_mc_constantin/fib_7_5_7_gt.h5', 'data')

    for id_ds_name, ds_name in enumerate(ds_names):
        for id_result_file, result_file in enumerate(result_files):
            result_key = result_keys[id_result_file]
            roi_and_rand_general(
                ds_name,
                project_folder,
                result_file,
                result_key,
                caching=True,
                debug=False,
                gt=gt,
                compute_rand=False
            )


def run_path_eval_to_xlsx():
    project_folder = '/mnt/localdata0/jhennies/results/multicut_workflow/170808_neuroproof/'
    thresh_range = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
    ds_names = ['fib_8_5_6', 'fib_8_5_7', 'fib_7_5_6']

    run_name = ''

    from eval_all import all_ds_path_eval
    results_path, results_obj = all_ds_path_eval(
        project_folder,
        thresh_range,
        ds_names,
        measures=None,
        run_name=''
    )

    print 'F1 = {}'.format(results_obj['f1'])
    print 'Recall = {}'.format(results_obj['recall'])
    print 'Precision = {}'.format(results_obj['precision'])
    print 'Accuracy = {}'.format(results_obj['accuracy'])

    error_measure_results_to_xlsx_file(
        os.path.join(project_folder, 'results_temp.xlsx'), results_path,
        np.array(thresh_range), ds_names,
        row_offset=0, append=False, title='Path level evaluation'
    )

    error_measure_results_to_xlsx_file(
        os.path.join(project_folder, 'results_temp.xlsx'), results_obj,
        np.array(thresh_range), ds_names,
        row_offset=len(thresh_range) + 5, append=True, title='Object level evaluation'
    )


if __name__ == '__main__':
    run_roi_and_rand_general()

    # run_path_eval_on_all_samples()
    # run_eval_obj_measures_readable()

    # run_plot_all_sample_path_eval_split_samples()





