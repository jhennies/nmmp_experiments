
import numpy as np
import os


def run_roi_and_rand_general():

    from eval_all import roi_and_rand_general

    project_folder = '/mnt/localdata1/jhennies/neuraldata/results/multicut_workflow/170606_branch_nifty_backend/'
    # project_folder = '/media/hdb/jhennies/neuraldata/results/multicut_workflow/170530_new_baseline/'
    # project_folder = '/media/julian/Daten/datasets/results/multicut_workflow/170530_new_baseline/'
    # result_file = 'result_resolved_local.h5'
    # result_files = ['result.h5', 'result_resolved_local.h5']
    result_files = ['result.h5', 'result_resolved_local.h5']
    source_folder = '/mnt/ssd/jhennies/neuraldata/cremi_2016/170606_resolve_false_merges/'
    # source_folder = '/media/julian/Daten/datasets/cremi_2016/170321_resolve_false_merges/'

    samples = ['A', 'A', 'B', 'B', 'C', 'C']
    halves = [0, 1] * 3
    defect_corrects = [False, False, True, True, True, True]

    # samples = ['A']
    # halves = [0]
    # defect_corrects = [False]

    # samples = ['A', 'A', 'B', 'B']
    # halves = [0, 1] * 2
    # defect_corrects = [False, False, True, True]

    for idx, sample in enumerate(samples):
        half = halves[idx]
        defect_correct = defect_corrects[idx]
        for result_file in result_files:
            roi_and_rand_general(sample, half, defect_correct, project_folder,
                                 source_folder, result_file, caching=True, debug=False)


def run_eval_obj_measures_readable():

    from eval_all import eval_obj_measures_readable

    samples = ['A', 'A', 'B', 'B', 'C', 'C']
    halves = [0, 1] * 4
    defect_corrects = [False, False, True, True, True, True]

    for spl_id, spl in enumerate(samples):

        half = halves[spl_id]
        defect_correct = defect_corrects[spl_id]

        print '\nEvaluating spl{}_z{}'.format(spl, half)

        project_folder = '/home/julian/ssh_data/neuraldata/results/multicut_workflow/170530_new_baseline/'
        seg_file = 'result.h5'
        seg_key = 'z/{}/data'.format(half)
        resolved_files = ['result_resolved_local.h5']
        resolved_key = seg_key

        result = eval_obj_measures_readable(
            spl, half,
            project_folder,
            seg_file, seg_key,
            resolved_files, resolved_key,
            thresh_range=None,
            resolved_only=False,
            defect_correct=defect_correct
        )
        result = result[0][0]

        # print result

        print 'Objects:            {}'.format(result[0])
        print 'TP:                 {}'.format(result[1])
        print '    Fully resolved: {}'.format(result[3])
        print '    Falsely split:  {}'.format(result[4])
        print 'FP:                 {}'.format(result[2])
        print '    Falsely split:  {}'.format(result[5])


def run_plot_all_sample_path_eval_split_samples():

    from eval_all import plot_all_sample_path_eval_split_samples
    samples = ['A', 'A', 'B', 'B', 'C', 'C']
    halves = [0, 1] * 3
    defect_corrects = [False, False, True, True, True, True]

    project_folder = '/home/julian/ssh_data/neuraldata/results/multicut_workflow/170530_new_baseline/'

    thresh_range = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]

    # Plot sample path evaluation
    for i in xrange(0, 6, 2):
        plot_all_sample_path_eval_split_samples(project_folder, thresh_range,
                                  halves[i: i+2],
                                  defect_corrects[i: i+2],
                                  samples[i: i+2],
                                  measures=['f1', 'precision', 'recall'])
    # plot_all_sample_path_eval_split_samples(
    #     project_folder,
    #     [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9],
    #     halves, defect_corrects,
    #     samples, measures=['f1', 'precision', 'recall']
    # )


def run_path_eval_on_all_samples():

    from eval_all import all_sample_path_eval

    project_folder = '/mnt/localdata1/jhennies/neuraldata/results/multicut_workflow/170622_new_path_features/'
    # thresh_range = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
    thresh_range = [0.3]
    samples = ['A', 'A', 'B', 'B', 'C', 'C']
    halves = [0, 1] * 3
    defect_corrects = [False, False, True, True, True, True]

    results_path, results_obj = all_sample_path_eval(
        project_folder,
        thresh_range,
        samples,
        halves,
        defect_corrects,
        measures=['f1', 'recall', 'precision']
    )

    print results_path
    print results_obj

    pass


def run_eval_obj_measures_readable():

    from eval_all import eval_obj_measures_readable

    spl = 'A'
    half = 0
    project_folder = '/mnt/localdata1/jhennies/neuraldata/results/multicut_workflow/170606_branch_nifty_backend/'
    seg_file = 'result.h5'
    seg_key = 'z/0/data'
    resolved_files = ['result_resolved_local.h5']
    resolved_keys = 'z/0/data'
    thresh_range = [0.3]

    result = eval_obj_measures_readable(
        spl,
        half,
        project_folder,
        seg_file,
        seg_key,
        resolved_files,
        resolved_keys,
        thresh_range,
        resolved_only=True,
        defect_correct=False
    )

    print result

    pass


def run_plot_all_sample_path_eval_split_samples():

    from eval_all import plot_all_sample_path_eval_split_samples

    project_folder = '/mnt/localdata1/jhennies/neuraldata/results/multicut_workflow/170622_new_path_features/'
    thresh_range = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
    # thresh_range = [0.3]
    # samples = ['A', 'A', 'B', 'B', 'C', 'C']
    # halves = [0, 1] * 3
    samples = ['A']
    halves = [0]
    defect_corrects = [False, False, True, True, True, True]

    plot_all_sample_path_eval_split_samples(
        project_folder,
        thresh_range,
        halves,
        defect_corrects,
        samples,
        measures=['f1', 'precision', 'recall']
    )


def error_measure_results_to_xlsx_file(filepath, results, thresh_range, samples, halves,
                                       row_offset=0, append=False, title=None):

    def write_list(ws, r, c, l):
        l = l.squeeze()
        assert l.ndim == 1
        # FIXME Is there a simpler command ommiting the for loop to write a list?
        for id, item in enumerate(l):
            ws.cell(row=r + id, column=c).value = item

    from openpyxl import Workbook, load_workbook

    if append:
        workbook = load_workbook(filepath)
    else:
        workbook = Workbook()

    worksheet = workbook.active

    if title is not None:
        worksheet.cell(row=1 + row_offset, column=1).value=title
        row_offset += 1

    # worksheet.cell(row=3, column=1).value = thresh_range
    write_list(worksheet, row_offset + 3, 1, thresh_range)

    for spl_id, spl in enumerate(samples):

        half = halves[spl_id]
        spl_str = 'Sample_{}{}'.format(spl, half)

        worksheet.cell(row=1 + row_offset, column=2 + spl_id * (len(results.keys()) + 1)).value = spl_str

        for measure_id, measure in enumerate(results.keys()):

            col = 2 + spl_id * (len(results.keys()) + 1) + measure_id

            worksheet.cell(row=2 + row_offset, column=col).value = measure
            write_list(worksheet, 3 + row_offset, col, results[measure][spl_id])

    workbook.save(filepath)


if __name__ == '__main__':
    # run_roi_and_rand_general()

    # run_path_eval_on_all_samples()
    # run_eval_obj_measures_readable()

    # run_plot_all_sample_path_eval_split_samples()

    project_folder = '/mnt/localdata1/jhennies/neuraldata/results/multicut_workflow/170717_new_path_features/'
    thresh_range = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
    samples = ['A', 'A', 'B', 'B', 'C', 'C']
    halves = [0, 1] * 3
    defect_corrects = [False, False, True, True, True, True]

    from eval_all import all_sample_path_eval
    results_path, results_obj = all_sample_path_eval(
        project_folder, thresh_range, samples, halves, defect_corrects,
        measures=['f1', 'precision', 'recall', 'accuracy'],
        source_folder='/mnt/ssd/jhennies/neuraldata/cremi_2016/170606_resolve_false_merges/'
    )

    print 'F1 = {}'.format(results_obj['f1'])
    print 'Recall = {}'.format(results_obj['recall'])
    print 'Precision = {}'.format(results_obj['precision'])
    print 'Accuracy = {}'.format(results_obj['accuracy'])

    error_measure_results_to_xlsx_file(
        os.path.join(project_folder, 'results_temp.xlsx'), results_path,
        np.array(thresh_range), samples, halves,
        row_offset=0, append=False, title='Path level evaluation'
    )

    error_measure_results_to_xlsx_file(
        os.path.join(project_folder, 'results_temp.xlsx'), results_obj,
        np.array(thresh_range), samples, halves,
        row_offset=len(thresh_range) + 5, append=True, title='Object level evaluation'
    )



